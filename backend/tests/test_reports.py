"""Excel report generation — the sheets actually build, and they respect scope.

653 lines of openpyxl had no test at all. Nothing here checks formatting; the
point is that every report type still produces a workbook, that its rows come
from the caller's own scope, and that the numbers agree with the analytics the
same user sees on screen — a report that quietly exports another campus's
tickets is the failure that matters, and it looks identical to a working one
from the outside.
"""

from io import BytesIO

import openpyxl
import pytest
from django.urls import reverse

from apps.tickets.services.lifecycle import transition_status
from tests import factories

pytestmark = pytest.mark.django_db

REPORT_IDS = [
    "ticket-lifecycle",
    "technician-performance",
    "facility-health",
    "pending-analysis",
    "comprehensive",
]


def _workbook(response):
    assert response.status_code == 200, response.content[:400]
    assert response["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    return openpyxl.load_workbook(BytesIO(response.content))


def _generate(api, report_type):
    return api.get(reverse("analytics:report-generate"), {"report_type": report_type})


@pytest.fixture
def worked_ticket(nrb_electrical_ticket, nrb_electrician, nrb_hos):
    """One ticket carried through to resolved, so the sheets have real rows:
    a resolution time to average, a pause to report, a rating to summarise."""
    t = nrb_electrical_ticket
    t.assigned_to = nrb_electrician
    t.save(update_fields=["assigned_to"])
    transition_status(t, "assigned", nrb_hos)
    transition_status(t, "in_progress", nrb_electrician)
    transition_status(t, "pending", nrb_electrician, pending_reason="awaiting_materials")
    transition_status(t, "in_progress", nrb_electrician)
    transition_status(t, "resolved", nrb_electrician)
    return t


@pytest.mark.parametrize("report_type", REPORT_IDS)
def test_every_report_type_builds_a_workbook(api, nrb_hos, worked_ticket, report_type):
    api.force_authenticate(nrb_hos)
    wb = _workbook(_generate(api, report_type))
    assert "Summary" in wb.sheetnames
    # Summary plus at least one data sheet; comprehensive carries all four.
    assert len(wb.sheetnames) >= 2
    if report_type == "comprehensive":
        assert len(wb.sheetnames) == 5


def test_the_lifecycle_sheet_lists_the_ticket(api, nrb_hos, worked_ticket):
    api.force_authenticate(nrb_hos)
    wb = _workbook(_generate(api, "ticket-lifecycle"))
    ws = wb[[n for n in wb.sheetnames if n != "Summary"][0]]
    values = {str(c.value) for row in ws.iter_rows() for c in row}
    assert worked_ticket.ticket_no in values


def test_a_report_never_exports_another_campus(
    api, msa_hos, worked_ticket, msa_electrical_ticket
):
    """The negative that makes the rest worth having. A HOS at Mombasa exports
    their own section; the Nairobi ticket must be absent, not merely unlabelled."""
    api.force_authenticate(msa_hos)
    wb = _workbook(_generate(api, "ticket-lifecycle"))
    ws = wb[[n for n in wb.sheetnames if n != "Summary"][0]]
    values = {str(c.value) for row in ws.iter_rows() for c in row}
    assert msa_electrical_ticket.ticket_no in values
    assert worked_ticket.ticket_no not in values


def test_report_type_is_required_and_validated(api, nrb_hos):
    api.force_authenticate(nrb_hos)
    assert api.get(reverse("analytics:report-generate")).status_code == 400
    assert _generate(api, "not-a-report").status_code == 400


def test_the_summary_sheet_agrees_with_the_analytics_endpoint(
    api, nrb_hos, worked_ticket, nrb_plumbing_ticket
):
    """The docstring on report_views promises the Summary mirrors the overview
    cards. If the two ever compute their own answer, this is what notices."""
    api.force_authenticate(nrb_hos)
    wb = _workbook(_generate(api, "ticket-lifecycle"))
    summary = wb["Summary"]
    cells = {}
    for row in summary.iter_rows(values_only=True):
        if row and row[0] is not None and len(row) > 1:
            cells[str(row[0]).strip().lower()] = row[1]

    headline = api.get(reverse("analytics:analytics")).json()["headline"]
    assert cells["open backlog (live)"] == headline["open_backlog"]
    assert cells["created in window"] == headline["created"]
    assert cells["resolved in window"] == headline["resolved"]
    assert cells["net flow (created \u2212 resolved)"] == headline["net_flow"]
